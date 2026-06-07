from collections import defaultdict
from csv import DictReader
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


BASE_DIR = Path(__file__).resolve().parent
DATA_FILE = BASE_DIR / "data" / "dummy_sales_data.csv"
OUTPUT_DIR = BASE_DIR / "outputs"
EXCEL_FILE = OUTPUT_DIR / "business_report.xlsx"
PDF_FILE = OUTPUT_DIR / "business_report.pdf"


def money(value):
    return f"${value:,.2f}"


def load_sales_rows():
    rows = []
    with DATA_FILE.open("r", encoding="utf-8", newline="") as file:
        reader = DictReader(file)
        for row in reader:
            units = int(row["units"])
            unit_price = float(row["unit_price"])
            order_date = datetime.strptime(row["order_date"], "%Y-%m-%d").date()
            rows.append(
                {
                    "order_id": row["order_id"],
                    "order_date": order_date,
                    "month": order_date.strftime("%Y-%m"),
                    "customer": row["customer"],
                    "region": row["region"],
                    "category": row["category"],
                    "product": row["product"],
                    "units": units,
                    "unit_price": unit_price,
                    "revenue": units * unit_price,
                }
            )
    return rows


def summarize(rows):
    revenue_by_region = defaultdict(float)
    revenue_by_category = defaultdict(float)
    revenue_by_month = defaultdict(float)
    revenue_by_customer = defaultdict(float)

    total_revenue = 0
    total_units = 0
    for row in rows:
        revenue = row["revenue"]
        total_revenue += revenue
        total_units += row["units"]
        revenue_by_region[row["region"]] += revenue
        revenue_by_category[row["category"]] += revenue
        revenue_by_month[row["month"]] += revenue
        revenue_by_customer[row["customer"]] += revenue

    average_order_value = total_revenue / len(rows) if rows else 0

    return {
        "total_orders": len(rows),
        "total_revenue": total_revenue,
        "total_units": total_units,
        "average_order_value": average_order_value,
        "revenue_by_region": dict(sorted(revenue_by_region.items())),
        "revenue_by_category": dict(sorted(revenue_by_category.items())),
        "revenue_by_month": dict(sorted(revenue_by_month.items())),
        "top_customers": dict(
            sorted(revenue_by_customer.items(), key=lambda item: item[1], reverse=True)[:10]
        ),
    }


def style_header(row):
    for cell in row:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")


def autosize_columns(sheet):
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        column_letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[column_letter].width = min(max_length + 3, 35)


def write_key_metrics(sheet, summary):
    sheet.title = "Dashboard"
    sheet.append(["Metric", "Value"])
    sheet.append(["Total Orders", summary["total_orders"]])
    sheet.append(["Total Revenue", summary["total_revenue"]])
    sheet.append(["Total Units Sold", summary["total_units"]])
    sheet.append(["Average Order Value", summary["average_order_value"]])
    style_header(sheet[1])
    for row_number in range(3, 6):
        sheet[f"B{row_number}"].number_format = "$#,##0.00"
    autosize_columns(sheet)


def add_summary_sheet(workbook, title, headers, rows):
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    style_header(sheet[1])
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
        row[0].number_format = "$#,##0.00"
    autosize_columns(sheet)
    return sheet


def add_bar_chart(sheet, title, anchor):
    chart = BarChart()
    chart.title = title
    chart.y_axis.title = "Revenue"
    chart.x_axis.title = sheet.cell(row=1, column=1).value
    data = Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 8
    chart.width = 14
    sheet.add_chart(chart, anchor)


def add_line_chart(sheet, title, anchor):
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = "Revenue"
    chart.x_axis.title = "Month"
    data = Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 8
    chart.width = 14
    sheet.add_chart(chart, anchor)


def create_excel_report(rows, summary):
    workbook = Workbook()
    dashboard = workbook.active
    write_key_metrics(dashboard, summary)

    raw_sheet = workbook.create_sheet("Raw Data")
    raw_headers = [
        "Order ID",
        "Order Date",
        "Customer",
        "Region",
        "Category",
        "Product",
        "Units",
        "Unit Price",
        "Revenue",
    ]
    raw_sheet.append(raw_headers)
    for row in rows:
        raw_sheet.append(
            [
                row["order_id"],
                row["order_date"].isoformat(),
                row["customer"],
                row["region"],
                row["category"],
                row["product"],
                row["units"],
                row["unit_price"],
                row["revenue"],
            ]
        )
    style_header(raw_sheet[1])
    for row in raw_sheet.iter_rows(min_row=2, min_col=8, max_col=9):
        for cell in row:
            cell.number_format = "$#,##0.00"
    autosize_columns(raw_sheet)

    region_sheet = add_summary_sheet(
        workbook,
        "Revenue by Region",
        ["Region", "Revenue"],
        summary["revenue_by_region"].items(),
    )
    category_sheet = add_summary_sheet(
        workbook,
        "Revenue by Category",
        ["Category", "Revenue"],
        summary["revenue_by_category"].items(),
    )
    month_sheet = add_summary_sheet(
        workbook,
        "Monthly Revenue",
        ["Month", "Revenue"],
        summary["revenue_by_month"].items(),
    )
    customer_sheet = add_summary_sheet(
        workbook,
        "Top Customers",
        ["Customer", "Revenue"],
        summary["top_customers"].items(),
    )

    add_bar_chart(region_sheet, "Revenue by Region", "D2")
    add_bar_chart(category_sheet, "Revenue by Category", "D2")
    add_line_chart(month_sheet, "Monthly Revenue Trend", "D2")
    add_bar_chart(customer_sheet, "Top Customers by Revenue", "D2")

    OUTPUT_DIR.mkdir(exist_ok=True)
    workbook.save(EXCEL_FILE)


def table_data_from_mapping(first_header, second_header, mapping):
    return [[first_header, second_header]] + [[key, money(value)] for key, value in mapping.items()]


def add_pdf_table(elements, title, data, styles):
    elements.append(Paragraph(title, styles["Heading2"]))
    table = Table(data, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9E2F3")),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F7FBFF")),
                ("PADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )
    elements.append(table)
    elements.append(Spacer(1, 14))


def create_pdf_report(summary):
    OUTPUT_DIR.mkdir(exist_ok=True)
    document = SimpleDocTemplate(str(PDF_FILE), pagesize=A4)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Automated Business Report", styles["Title"]),
        Paragraph("Generated from dummy raw sales data", styles["Normal"]),
        Spacer(1, 16),
    ]

    metrics = [
        ["Metric", "Value"],
        ["Total Orders", summary["total_orders"]],
        ["Total Revenue", money(summary["total_revenue"])],
        ["Total Units Sold", summary["total_units"]],
        ["Average Order Value", money(summary["average_order_value"])],
    ]
    add_pdf_table(elements, "Key Metrics", metrics, styles)
    add_pdf_table(
        elements,
        "Revenue by Region",
        table_data_from_mapping("Region", "Revenue", summary["revenue_by_region"]),
        styles,
    )
    add_pdf_table(
        elements,
        "Revenue by Category",
        table_data_from_mapping("Category", "Revenue", summary["revenue_by_category"]),
        styles,
    )
    add_pdf_table(
        elements,
        "Monthly Revenue",
        table_data_from_mapping("Month", "Revenue", summary["revenue_by_month"]),
        styles,
    )
    add_pdf_table(
        elements,
        "Top Customers",
        table_data_from_mapping("Customer", "Revenue", summary["top_customers"]),
        styles,
    )

    document.build(elements)


def main():
    rows = load_sales_rows()
    summary = summarize(rows)
    create_excel_report(rows, summary)
    create_pdf_report(summary)
    print(f"Excel report created: {EXCEL_FILE}")
    print(f"PDF report created: {PDF_FILE}")


if __name__ == "__main__":
    main()

